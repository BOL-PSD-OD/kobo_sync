"""Read a Kobo export, load prior IDs from an existing master, write master.xlsx."""

import csv
import os

import openpyxl

MASTER_COLUMNS = ["store_id", "status", "shop_name", "biz_type", "district",
                  "is_duplicate", "is_catalog", "uuid"]


def read_export(path):
    """Read a Kobo export (.xlsx or .csv). Return (columns, rows) where rows is
    a list of dicts colname->value."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            cols = list(reader.fieldnames or [])
            return cols, [dict(r) for r in reader]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    cols = ["" if v is None else str(v) for v in next(it)]
    rows = []
    for r in it:
        if all(v is None for v in r):
            continue
        rows.append({cols[i]: r[i] for i in range(len(cols))})
    return cols, rows


def load_prior_ids(master_path):
    """Return {uuid: store_id} from an existing master.xlsx, or {} if absent."""
    if not os.path.exists(master_path):
        return {}
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = ["" if v is None else str(v) for v in next(it)]
    ix = {h: i for i, h in enumerate(header)}
    out = {}
    if "uuid" not in ix or "store_id" not in ix:
        return out
    for r in it:
        uuid = r[ix["uuid"]]
        sid = r[ix["store_id"]]
        if uuid and sid:
            out[str(uuid)] = str(sid)
    return out


def write_master(master_path, rows):
    """Write master rows to xlsx, overwriting. Creates parent dir if needed."""
    os.makedirs(os.path.dirname(master_path), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "master"
    ws.append(MASTER_COLUMNS)
    for row in rows:
        ws.append([row.get(c, "") for c in MASTER_COLUMNS])
    wb.save(master_path)
